from fastapi import status, Depends, APIRouter, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from .. import models, schemas, oauth2
from ..database import get_db
from ..models import User
from typing import Optional, Annotated
from openpyxl import Workbook, load_workbook
from sqlalchemy import inspect
from io import BytesIO
from fastapi.responses import StreamingResponse
from tempfile import NamedTemporaryFile
from datetime import datetime

router = APIRouter(
    prefix="/v1/report",
    tags=["Report"]
)


@router.get("/", response_model=None)
def get_users(limit: Optional[int] = 10, skip: Optional[int] = 0, search: Optional[str] = "", 
              sort: Optional[str] = "id", order: Optional[str] = "desc",
              db: Session = Depends(get_db), current_user: User = Depends(oauth2.get_current_user)):
    users = db.query(models.User)
    if search:
         users = users.filter(models.User.email.contains(search))
    if limit:
        users = users.limit(limit).offset(skip)
    users = users.all()
    # Create excel file
    wb = Workbook()
    ws = wb.active
    mapper = inspect(models.User)
    headers =[column.name for column in mapper.columns]
    headers.remove("password")
    ws.append(headers)
    
    for row_data in users:
        ws.append([getattr(row_data, key) for key in headers])
    # Iterate over all columns and adjust their widths
    for column in ws.columns:
        max_length = 15
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width
    # wb.save(filename='data.xlsx') -> save file
    with NamedTemporaryFile() as tmp:
     wb.save(tmp.name)
     data = BytesIO(tmp.read())

    file_name =str(round(datetime.now().timestamp()))+'_'"users"

    return StreamingResponse(data, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={
                "Content-Disposition": f'attachment; filename="{file_name}.xlsx"'
    })

@router.post("/users", status_code=status.HTTP_201_CREATED)
def import_users(file: Annotated[UploadFile, File(description="A file read as UploadFile")], db: Session = Depends(get_db),
              current_user: User = Depends(oauth2.get_current_user)):
    if not file:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No file sent")
    if file.content_type != 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
        raise HTTPException(400, detail="Invalid document type")
    # Read file
    workbook = load_workbook(BytesIO(file.file.read()))
    sheet_obj = workbook.active
    people = []
    default_pass = 123456    
    for row in list(sheet_obj.rows)[1:]:
        args = [cell.value for cell in row]
        people.append(args)
    # Data
    print(people)
    # Process ...
    return {"message": "Successfully upload file"}